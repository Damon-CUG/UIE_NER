# -*- coding: utf-8 -*-
# coding:unicode_escape
"""
Created on Mon Jan  6 20:47:37 2020

@author: cm
"""

import re
import csv
import sys
from openpyxl import load_workbook

class ToolGeneral():
    """
    Tool function
    """
    def is_odd(self,num):
        if num % 2 == 0:
            return 'even'
        else:
            return 'odd'        

    def load_dict(self,file):  
        """
        Load dictionary
        """
        with  open(file,encoding='utf-8', errors='ignore') as fp:
            lines = fp.readlines()
            lines = [l.strip() for l in lines]
            print("Load data from file (%s) finished !"%file)
            dictionary = [word.strip() for word in lines]
        return set(dictionary)

    def load_emotion_dict(self,file):
        """
        Load emotion dictionary
        """
        dictionary = {}
        wb = load_workbook(file)
        sheets = wb.worksheets  # 获取当前所有的sheet
        print(sheets)

        # 获取第一张sheet
        sheet1 = sheets[0]
        #print(sheet1.max_row)
        for i in range(1, sheet1.max_row + 1):
# <<<<<<< HEAD
#             if sheet1.cell(i, 1).value != '词语':
#                 key = sheet1.cell(i, 1).value
#                 value = sheet1.cell(i, 2).value
# =======
            if sheet1.cell(row=i, column=1).value != '词语':
                key = sheet1.cell(row=i, column=1).value
                value = sheet1.cell(row=i, column=2).value
# >>>>>>> dev
                dictionary[key] = value

        return dictionary



    def sentence_split_regex(self,sentence):
        """
        Segmentation of sentence
        """
        if sentence is not None:
            sentence = re.sub(r"&ndash;+|&mdash;+", "-", sentence)
            sub_sentence = re.split(r"[。,，！!？?;；\s…~～]+|\.{2,}|&hellip;+|&nbsp+|_n|_t", sentence)
            sub_sentence = [s for s in sub_sentence if s != '']
            if sub_sentence != []:
                return sub_sentence
            else:
                return [sentence]
        return []
    
    
if __name__ == "__main__": 
    #
    tool = ToolGeneral()    
    #
    s = '我今天。昨天上午，还有现在'
    ls = tool.sentence_split_regex(s)
    dict = tool.load_emotion_dict('./label/angry.xlsx')
    print(dict)
    
    
    
